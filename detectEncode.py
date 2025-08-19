import pandas as pd

def csv_to_xlsx_all_text(csv_file, xlsx_file):
    # Read CSV, force all values as string
    df = pd.read_csv(csv_file, dtype=str)

    # Write to Excel
    df.to_excel(xlsx_file, index=False)

# Example
csv_to_xlsx_all_text("input.csv", "output.xlsx")
print("Done! All columns saved as text")


from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# 打开或新建工作簿
try:
    wb = load_workbook("test.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active

# 条件格式：如果 B2 <> A2，则 B2 背景色变红
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

rule = FormulaRule(formula=['$B2<>$A2'], fill=red_fill)

# 给 B2 单元格应用规则
ws.conditional_formatting.add("B2", rule)

wb.save("test.xlsx")
